# coding=utf-8

"""
Data pre-processing for PES dataset
https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Goalkeeper.xlsx
https://github.com/ShenbaoYu/TCFPACN/blob/002649cfd8f67d9678cdc36564a1bc024fb87b7f/Data/PES/Goalkeeper.xlsx

https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Back.xlsx
https://github.com/ShenbaoYu/TCFPACN/blob/002649cfd8f67d9678cdc36564a1bc024fb87b7f/Data/PES/Back.xlsx

https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Forward.xlsx
https://github.com/ShenbaoYu/TCFPACN/blob/002649cfd8f67d9678cdc36564a1bc024fb87b7f/Data/PES/Forward.xlsx

"""

import os, sys
BASE_DIR = os.path.dirname(os.path.dirname(os.getcwd()))
sys.path.append(BASE_DIR)
sys.path.append('TCFPACN')
# A leitura de dados do arquivo XLSX é feita usando a biblioteca openpyxl,
# que é específica para trabalhar com arquivos Excel.
import openpyxl
import math
from FBTP import players
from collections import OrderedDict

import requests

urls = [
    'https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Goalkeeper.xlsx',
    'https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Back.xlsx',
    'https://github.com/ShenbaoYu/TCFPACN/raw/main/Data/PES/Forward.xlsx',
]  # Substitua pelos links corretos

for url in urls:
    filename = url.split('/')[-1]  # Extrai o nome do arquivo da URL
    response = requests.get(url)
    with open(filename, 'wb') as f:
        f.write(response.content)


"""
Idêntica à função do código FIFApre, mas para o conjunto de dados PES.
Lê os critérios de avaliação de um arquivo de texto e normaliza seus valores.
"""
def read_criteria(path, criteria_file):
    criteria = {}
    with open(path + criteria_file, 'r') as cf:
        while True:
            line = cf.readline()
            if not line:
                break
            name = line.split(":")[0]
            value = line.split(":")[1][:-1]

            criteria[name] = int(value)

    criteria = normalize(criteria)

    return criteria

"""
- Lê informações sobre goleiros de um arquivo XLSX.
- Cria objetos Goalkeeper para cada goleiro, preenchendo seus atributos (ID, rating, salário, habilidades).
- O salário é calculado da mesma forma que no código 3, com base no rating do goleiro.
"""
def get_goalkeepers(file_name):

    wb = openpyxl.load_workbook(file_name)
    #wb = openpyxl.load_workbook(url)   # open source
    ws = wb["GoalKeeper"]

    goal_keepers = []
    for row in range(2, ws.max_row+1):
        gk_id = ws.cell(row=row, column=1).value
        gk = players.Goalkeeper(gk_id)
        gk.rating = ws.cell(row=row, column=10).value  # the rating of goalkeeprs
        gk.salary = 0.0006375 * math.exp(0.1029*gk.rating)  # calculate salary

        for column in range(29, ws.max_column+1): # read skills
            gk.ability.append(ws.cell(row=row, column=column).value)
        goal_keepers.append(gk)

    return goal_keepers

"""
- Lê informações sobre jogadores (exceto goleiros) de um arquivo XLSX.
- Cria dicionários para armazenar:
--> Atributos dos jogadores (nome do time, nacionalidade).
--> Habilidades dos jogadores (com seus respectivos valores).
--> Posições dos jogadores.
--> Ratings dos jogadores.
--> IDs dos jogadores.
- Preenche esses dicionários com os dados lidos do arquivo XLSX.

 OBS: Não possui a lógica para lidar com jogadores sem posição definida,
 como no <FIFApre>. Isso pode ser uma diferença nos dados do PES ou uma escolha de implementação.

"""
def read_info(file_name):

    ability_name_id = {}  # ability {name:id}
    player_attributes = {}  # atributos do jogador, incluindo clube e nacionalidade
    player_abilities_name = OrderedDict()  # habilidades pessoais
    player_position = {}  # posição do jogador {id:position}
    player_rating = {}  # avaliação do jogador {id:rating}

    wb = openpyxl.load_workbook(file_name)
    #wb = openpyxl.load_workbook(path + file_name)
    ws = wb["Players"]

    player_no_id = {}  # player's number : id
    no = 0  # number of player
    ability_id = 0  # número da habilidade

    for row in range(1, ws.max_row + 1):
        if row == 1:
            for column in range(11, 29):  # target ability
                player_abilities_name[ws.cell(row=row, column=column).value] = {}

            for ability_name in player_abilities_name.keys():
                ability_name_id[ability_name] = ability_id
                ability_id += 1
        else:
            player_id = ws.cell(row=row, column=1).value  # get player's ID
            player_no_id[no] = player_id
            position = ws.cell(row=row, column=2).value  # get player's position
            player_position[no] = position
            rating = ws.cell(row=row, column=10).value  # get player's rating
            player_rating[no] = rating

            # player's attributes, including club and nationality
            team_name = ws.cell(row=row, column=4).value  # player's team name
            nationality = ws.cell(row=row, column=5).value  # player's nationality
            player_attributes[no] = []
            player_attributes[no].append(team_name)
            player_attributes[no].append(nationality)

            # player's abilities
            _ = []
            for column in range(11, 29):  # range of target abilities
                _.append(ws.cell(row=row, column=column).value)
            c = 0
            for ability in player_abilities_name:
                player_abilities_name[ability][no] = _[c]
                c += 1

            no += 1

    return ability_name_id,       \
           player_attributes,     \
           player_abilities_name, \
           player_position,       \
           player_rating,         \
           player_no_id

"""
- Função auxiliar usada por read_criteria para normalizar os valores de um dicionário.
- Divide cada valor pela soma total dos valores, garantindo que a soma dos valores normalizados seja 1.
"""
def normalize(dict_type):
    total = sum(v for v in dict_type.values())
    tmp = {}
    for key, value in dict_type.items():
        tmp[key] = value / total
    return tmp

goalkeepers = get_goalkeepers('Goalkeeper.xlsx')
ability_name_id, player_attributes, player_abilities_name, player_position, player_rating, player_no_id = read_info('Back.xlsx')
ability_name_id, player_attributes, player_abilities_name, player_position, player_rating, player_no_id = read_info('Forward.xlsx')